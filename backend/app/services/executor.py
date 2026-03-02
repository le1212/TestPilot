import httpx
import time
import json
import re
import os
from datetime import datetime, timezone
from typing import Any


def _excel_rows(file_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """读取 Excel，首行为表头，返回 [{"col1": "v1", ...}, ...]。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if not os.path.isfile(file_path):
        return []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"_c{i}" for i, h in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        out.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    wb.close()
    return out


def _replace_row_variables(obj: Any, row: dict[str, Any]) -> Any:
    """将字符串或 dict 中的 {{列名}} 替换为 row 对应列的值。"""
    if obj is None:
        return None
    if isinstance(obj, str):
        s = obj
        for k, v in row.items():
            s = s.replace(f"{{{{{k}}}}}", str(v) if v is not None else "")
        return s
    if isinstance(obj, dict):
        return {key: _replace_row_variables(val, row) for key, val in obj.items()}
    if isinstance(obj, list):
        return [_replace_row_variables(item, row) for item in obj]
    return obj


class TestExecutor:
    """Unified test execution engine supporting API, Web, App, and Mini-program testing."""

    @staticmethod
    async def execute_api(config: dict, env_base_url: str = "", env_headers: dict = None, env_variables: dict = None) -> dict:
        data_driver = config.get("data_driver") or {}
        if data_driver.get("enabled") and data_driver.get("file_url"):
            return await TestExecutor._execute_api_data_driven(config, env_base_url, env_headers, env_variables, data_driver)

        return await TestExecutor._execute_api_single(config, env_base_url, env_headers, env_variables)

    @staticmethod
    async def _execute_api_data_driven(
        config: dict, env_base_url: str, env_headers: dict | None, env_variables: dict | None, data_driver: dict
    ) -> dict:
        uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
        file_url = data_driver.get("file_url", "")
        filename = file_url.split("/")[-1].split("?")[0] if file_url else ""
        file_path = os.path.join(uploads_dir, filename) if filename else ""
        sheet_name = data_driver.get("sheet_name") or None
        rows = _excel_rows(file_path, sheet_name)
        if not rows:
            return {
                "result": {"passed": False, "response": {}, "assertions": [], "error": "Excel 无数据或文件不存在"},
                "logs": f"[数据驱动] 未读取到行数据: {file_path}",
                "duration_ms": 0,
            }
        all_logs = []
        total_duration = 0
        results = []
        for i, row in enumerate(rows):
            row_config = {**config}
            row_config["url"] = _replace_row_variables(config.get("url", ""), row)
            row_config["body"] = _replace_row_variables(config.get("body"), row)
            row_config["params"] = _replace_row_variables(config.get("params") or {}, row)
            row_config["headers"] = _replace_row_variables(config.get("headers") or {}, row)
            row_config.pop("data_driver", None)
            one = await TestExecutor._execute_api_single(row_config, env_base_url, env_headers, env_variables)
            all_logs.append(f"[数据驱动 第{i + 1}行] {one.get('logs', '')}")
            total_duration += one.get("duration_ms", 0)
            results.append(one)
        all_passed = all((r.get("result") or {}).get("passed", False) for r in results)
        return {
            "result": {
                "passed": all_passed,
                "response": {"data_driver_rows": len(results), "passed": sum(1 for r in results if (r.get("result") or {}).get("passed"))},
                "assertions": [],
                "error": None if all_passed else f"共 {len(results)} 行，{sum(1 for r in results if (r.get('result') or {}).get('passed'))} 通过",
            },
            "logs": "\n".join(all_logs),
            "duration_ms": total_duration,
        }

    @staticmethod
    async def _execute_api_single(config: dict, env_base_url: str = "", env_headers: dict = None, env_variables: dict = None) -> dict:
        url = config.get("url", "")
        method = config.get("method", "GET").upper()
        headers = {**(env_headers or {}), **config.get("headers", {})}
        params = config.get("params", {})
        body = config.get("body", None)
        body_type = config.get("body_type", "json")
        assertions = config.get("assertions", [])

        if env_variables:
            url = TestExecutor._replace_variables(url, env_variables)
            if body and isinstance(body, str):
                body = TestExecutor._replace_variables(body, env_variables)
            headers = {k: TestExecutor._replace_variables(v, env_variables) for k, v in headers.items()}
            params = {k: TestExecutor._replace_variables(str(v), env_variables) for k, v in params.items()}

        if env_base_url and not url.startswith("http"):
            url = env_base_url.rstrip("/") + "/" + url.lstrip("/")

        timeout_sec = float(config.get("timeout_seconds", 30))
        if timeout_sec <= 0 or timeout_sec > 300:
            timeout_sec = 30.0

        logs = []
        result = {"passed": False, "response": {}, "assertions": [], "error": None}

        start = time.time()
        try:
            async with httpx.AsyncClient(timeout=timeout_sec, verify=False) as client:
                logs.append(f"[请求] {method} {url}")
                logs.append(f"[请求] 请求头: {json.dumps(headers, ensure_ascii=False)}")
                if params:
                    logs.append(f"[请求] 参数: {json.dumps(params, ensure_ascii=False)}")

                kwargs: dict[str, Any] = {"headers": headers, "params": params}
                if method in ("POST", "PUT", "PATCH", "DELETE"):
                    if body_type == "json" and body:
                        parsed = json.loads(body) if isinstance(body, str) else body
                        kwargs["json"] = parsed
                        logs.append(f"[请求] 请求体(JSON): {json.dumps(parsed, ensure_ascii=False)}")
                    elif body_type == "form" and body:
                        kwargs["data"] = body if isinstance(body, dict) else json.loads(body)
                    elif body:
                        kwargs["content"] = body if isinstance(body, str) else json.dumps(body)

                resp = await client.request(method, url, **kwargs)

            duration = int((time.time() - start) * 1000)
            logs.append(f"[响应] 状态码: {resp.status_code} ({duration}ms)")

            try:
                resp_body = resp.json()
            except Exception:
                resp_body = resp.text

            logs.append(f"[响应] 响应内容: {json.dumps(resp_body, ensure_ascii=False)[:2000]}")

            result["response"] = {
                "status_code": resp.status_code,
                "headers": dict(resp.headers),
                "body": resp_body,
                "duration_ms": duration,
            }

            all_passed = True
            for assertion in assertions:
                a_result = TestExecutor._check_assertion(assertion, resp, resp_body)
                result["assertions"].append(a_result)
                if not a_result["passed"]:
                    all_passed = False
                    logs.append(f"[失败] {a_result['message']}")
                else:
                    logs.append(f"[通过] {a_result['message']}")

            if not assertions:
                all_passed = 200 <= resp.status_code < 400
                logs.append(f"[{'通过' if all_passed else '失败'}] 状态码断言: {resp.status_code}")

            result["passed"] = all_passed

        except Exception as e:
            duration = int((time.time() - start) * 1000)
            result["error"] = str(e)
            if "response" not in result or not isinstance(result["response"], dict):
                result["response"] = {}
            result["response"]["duration_ms"] = duration
            logs.append(f"[错误] {str(e)}")

        return {"result": result, "logs": "\n".join(logs), "duration_ms": result["response"].get("duration_ms", 0)}

    @staticmethod
    def _replace_variables(text: str, variables: dict) -> str:
        for key, value in variables.items():
            text = text.replace(f"{{{{{key}}}}}", str(value))
        return text

    @staticmethod
    def _check_assertion(assertion: dict, resp, resp_body) -> dict:
        a_type = assertion.get("type", "status_code")
        expected = assertion.get("expected")
        field = assertion.get("field", "")
        operator = assertion.get("operator", "equals")

        type_label = {
            "status_code": "状态码",
            "json_path": "JSON路径",
            "header": "响应头",
            "body_contains": "响应体包含",
            "response_time": "响应耗时",
        }.get(a_type, a_type)

        op_label = {
            "equals": "等于",
            "not_equals": "不等于",
            "contains": "包含",
            "not_contains": "不包含",
            "greater_than": "大于",
            "less_than": "小于",
            "regex": "正则匹配",
        }.get(operator, operator)

        try:
            if a_type == "status_code":
                actual = resp.status_code
            elif a_type == "json_path":
                actual = TestExecutor._extract_json_path(resp_body, field)
            elif a_type == "header":
                actual = resp.headers.get(field, "")
            elif a_type == "body_contains":
                actual = resp.text if hasattr(resp, "text") else str(resp_body)
                passed = str(expected) in actual
                return {"type": a_type, "field": field, "expected": expected, "actual": f"(长度: {len(actual)})", "passed": passed, "message": f"响应体{'包含' if passed else '不包含'} '{expected}'"}
            elif a_type == "response_time":
                actual = resp.elapsed.total_seconds() * 1000 if hasattr(resp, "elapsed") else 0
            else:
                return {"type": a_type, "passed": False, "message": f"未知断言类型: {a_type}"}

            passed = TestExecutor._compare(actual, expected, operator)
            return {
                "type": a_type,
                "field": field,
                "operator": operator,
                "expected": expected,
                "actual": actual,
                "passed": passed,
                "message": f"{type_label}({field}): {actual} {op_label} {expected} -> {'通过' if passed else '失败'}"
            }
        except Exception as e:
            return {"type": a_type, "passed": False, "message": f"断言执行异常: {str(e)}"}

    @staticmethod
    def _extract_json_path(data, path: str):
        keys = path.replace("[", ".").replace("]", "").split(".")
        current = data
        for key in keys:
            if not key:
                continue
            if isinstance(current, dict):
                current = current[key]
            elif isinstance(current, list):
                current = current[int(key)]
            else:
                raise ValueError(f"Cannot traverse '{key}' in {type(current)}")
        return current

    @staticmethod
    def _compare(actual, expected, operator: str) -> bool:
        if operator == "equals":
            return str(actual) == str(expected)
        elif operator == "not_equals":
            return str(actual) != str(expected)
        elif operator == "contains":
            return str(expected) in str(actual)
        elif operator == "not_contains":
            return str(expected) not in str(actual)
        elif operator == "greater_than":
            return float(actual) > float(expected)
        elif operator == "less_than":
            return float(actual) < float(expected)
        elif operator == "regex":
            return bool(re.search(str(expected), str(actual)))
        return False
