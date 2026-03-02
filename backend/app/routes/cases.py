from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional
from io import BytesIO
import json
from ..database import get_db
from ..db_utils import escape_like
from ..date_utils import parse_date_to_utc_range
from ..models import TestCase, User, CaseGroup, TestExecution, Defect, Project
from ..schemas import TestCaseCreate, TestCaseUpdate, TestCaseOut, TestCaseBatchUpdate
from ..user_utils import get_user_display_with_account
from .auth import get_current_user
from .projects import _can_access_project
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/cases", tags=["test-cases"])


def _can_delete_case(user: User, case: TestCase) -> bool:
    return bool(user.is_admin or (case.created_by_id and case.created_by_id == user.id))


def _can_edit_case(user: User, case: TestCase) -> bool:
    """创建人、管理员或协作者可编辑"""
    if not user:
        return False
    if user.is_admin or (case.created_by_id and case.created_by_id == user.id):
        return True
    cids = getattr(case, "collaborator_ids", None)
    if isinstance(cids, list) and user.id in cids:
        return True
    return False


def can_run_case(user: User, case: TestCase) -> bool:
    """仅创建人、管理员或协作者可执行用例。与编辑权限一致。"""
    return _can_edit_case(user, case)


def _can_manage_group(user: User, group: CaseGroup) -> bool:
    if not user:
        return False
    if user.is_admin or (group.created_by_id and group.created_by_id == user.id):
        return True
    cids = getattr(group, "collaborator_ids", None)
    return isinstance(cids, list) and user.id in cids


@router.get("", response_model=list[TestCaseOut])
async def list_cases(
    project_id: Optional[int] = Query(None),
    group_id: Optional[int] = Query(None),
    type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    tags: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if project_id:
        proj_result = await db.execute(select(Project).where(Project.id == project_id))
        project = proj_result.scalar_one_or_none()
        if not project:
            raise HTTPException(404, "项目不存在")
        if not _can_access_project(user, project):
            raise HTTPException(403, "无权限查看该项目下的用例")
    stmt = select(TestCase)
    if project_id:
        stmt = stmt.where(TestCase.project_id == project_id)
    if group_id is not None:
        stmt = stmt.where(TestCase.group_id == group_id)
    if type:
        stmt = stmt.where(TestCase.type == type)
    if priority:
        stmt = stmt.where(TestCase.priority == priority)
    if status:
        stmt = stmt.where(TestCase.status == status)
    if keyword:
        kw = escape_like(keyword.strip())
        stmt = stmt.where(TestCase.name.like(f"%{kw}%", escape="\\"))
    if tags:
        tag_list = [t.strip() for t in tags.split(",") if t.strip()]
        if tag_list:
            from sqlalchemy import or_, String
            stmt = stmt.where(
                or_(*[TestCase.tags.cast(String).contains(f'"{t}"') for t in tag_list])
            )
    r = parse_date_to_utc_range(date)
    if r:
        stmt = stmt.where(TestCase.updated_at >= r[0], TestCase.updated_at < r[1])

    stmt = stmt.order_by(TestCase.updated_at.desc())
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(stmt)
    cases = result.scalars().all()
    ids = [c.created_by_id for c in cases if getattr(c, "created_by_id", None)]
    names = await get_user_display_with_account(db, ids)

    proj_ids = list({c.project_id for c in cases if c.project_id})
    proj_map: dict[int, str] = {}
    if proj_ids:
        pr = await db.execute(select(Project.id, Project.name).where(Project.id.in_(proj_ids)))
        proj_map = {r[0]: r[1] for r in pr.all()}

    out = []
    for c in cases:
        o = TestCaseOut.model_validate(c)
        o.created_by_name = names.get(c.created_by_id, "") if getattr(c, "created_by_id", None) else None
        o.project_name = proj_map.get(c.project_id, "")
        out.append(o)
    return out


@router.get("/count")
async def count_cases(
    project_id: Optional[int] = Query(None),
    type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(func.count()).select_from(TestCase)
    if project_id:
        stmt = stmt.where(TestCase.project_id == project_id)
    if type:
        stmt = stmt.where(TestCase.type == type)
    result = await db.execute(stmt)
    return {"count": result.scalar() or 0}


# 导出用例为 Excel（与 list_cases 相同的筛选条件；建议传 group_id 导出指定分组，最多 10000 条）
_EXPORT_PAGE_SIZE = 10000
_TYPE_LABELS = {"api": "接口", "web": "网页", "app": "App", "miniapp": "小程序"}
_PRIORITY_LABELS = {"low": "低", "medium": "中", "high": "高", "critical": "严重"}
_STATUS_LABELS = {"draft": "草稿", "active": "启用", "deprecated": "废弃"}
_EXEC_STATUS_LABELS = {"pending": "待执行", "running": "执行中", "passed": "通过", "failed": "失败", "error": "错误", "cancelled": "已取消"}

# 断言类型中文
_ASSERT_TYPE_LABELS = {"status_code": "状态码", "json_path": "JSON路径", "header": "响应头", "body_contains": "包含文本"}


def _format_export_columns(config: dict, case_type: str) -> tuple:
    """返回 (操作步骤, 输入数据, 预期结果)。"""
    if not config:
        return ("", "", "")
    if case_type == "api":
        method = (config.get("method") or "GET").strip().upper()
        url = (config.get("url") or "").strip()
        steps = f"1. 发送 {method} 请求：{url}" if url else "发送请求"
        params = config.get("params") or {}
        config.get("body_type") or "json"
        body = config.get("body") or ""
        input_parts = []
        if isinstance(params, dict) and params:
            input_parts.append(json.dumps(params, ensure_ascii=False))
        if body:
            input_parts.append(body.strip()[:3000] + ("..." if len(body.strip()) > 3000 else ""))
        input_data = "\n".join(input_parts) if input_parts else ""
        assertions = config.get("assertions") or []
        expect_lines = []
        for a in assertions:
            if not isinstance(a, dict):
                continue
            t = a.get("type") or "status_code"
            label = _ASSERT_TYPE_LABELS.get(t, t)
            field = (a.get("field") or "").strip()
            op = a.get("operator") or "equals"
            expected = a.get("expected")
            if field:
                expect_lines.append(f"{label} {field} {op} {expected}")
            else:
                expect_lines.append(f"{label} {op} {expected}")
        expected_result = "\n".join(expect_lines)
        return (steps, input_data, expected_result)
    # web / app / miniapp
    steps_list = config.get("steps") or []
    step_lines = []
    input_parts = []
    expect_parts = []
    for i, s in enumerate(steps_list, 1):
        if not isinstance(s, dict):
            continue
        action = (s.get("action") or "").strip()
        if action.lower() == "__group__":
            step_lines.append(f"{i}. [分组] {s.get('value') or s.get('description') or ''}")
            continue
        locator = (s.get("locator") or "").strip()
        value = (s.get("value") or "").strip()
        desc = (s.get("description") or "").strip()
        step_lines.append(f"{i}. {action}" + (f" 定位:{locator}" if locator else "") + (f" 值:{value}" if value else "") + (f" — {desc}" if desc else ""))
        if value:
            input_parts.append(value[:500])
        if desc and ("断言" in desc or "验证" in desc or "预期" in desc or "assert" in action.lower()):
            expect_parts.append(desc)
    return ("\n".join(step_lines), "\n".join(input_parts), "\n".join(expect_parts))


@router.get("/export")
async def export_cases_excel(
    project_id: Optional[int] = Query(None),
    group_id: Optional[int] = Query(None),
    ungrouped_only: bool = Query(False, description="仅导出未分组用例"),
    type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    if not project_id:
        raise HTTPException(400, "请指定 project_id 再导出")
    stmt = select(TestCase).where(TestCase.project_id == project_id)
    if ungrouped_only:
        stmt = stmt.where(TestCase.group_id.is_(None))
    elif group_id is not None:
        stmt = stmt.where(TestCase.group_id == group_id)
    if type:
        stmt = stmt.where(TestCase.type == type)
    if priority:
        stmt = stmt.where(TestCase.priority == priority)
    if status:
        stmt = stmt.where(TestCase.status == status)
    if keyword:
        kw = escape_like(keyword.strip())
        stmt = stmt.where(TestCase.name.like(f"%{kw}%", escape="\\"))
    r = parse_date_to_utc_range(date)
    if r:
        stmt = stmt.where(TestCase.updated_at >= r[0], TestCase.updated_at < r[1])
    stmt = stmt.order_by(TestCase.group_id.asc(), TestCase.id.asc()).limit(_EXPORT_PAGE_SIZE)
    result = await db.execute(stmt)
    cases = result.scalars().all()
    case_ids = [c.id for c in cases]
    creator_ids = [c.created_by_id for c in cases if getattr(c, "created_by_id", None)]
    names = await get_user_display_with_account(db, creator_ids) if creator_ids else {}

    group_ids = [c.group_id for c in cases if c.group_id is not None]
    group_names = {}
    if group_ids:
        gr = await db.execute(select(CaseGroup).where(CaseGroup.id.in_(group_ids)))
        for g in gr.scalars().all():
            group_names[g.id] = g.name

    # 每个用例最近一次执行（按 finished_at 降序取第一条）
    last_execution_map = {}
    executor_name_map = {}
    if case_ids:
        ex_stmt = (
            select(TestExecution)
            .where(TestExecution.test_case_id.in_(case_ids))
            .order_by(TestExecution.id.desc())
        )
        ex_result = await db.execute(ex_stmt)
        all_executions = ex_result.scalars().all()
        seen = set()
        for ex in all_executions:
            if ex.test_case_id not in seen:
                seen.add(ex.test_case_id)
                last_execution_map[ex.test_case_id] = ex
        executor_ids = [ex.created_by_id for ex in last_execution_map.values() if getattr(ex, "created_by_id", None)]
        if executor_ids:
            executor_name_map = await get_user_display_with_account(db, executor_ids)

    # 每个用例关联的缺陷 ID 列表
    defect_map = {}
    if case_ids:
        def_stmt = select(Defect).where(Defect.test_case_id.in_(case_ids))
        def_result = await db.execute(def_stmt)
        for d in def_result.scalars().all():
            defect_map.setdefault(d.test_case_id, []).append(d.id)

    headers = [
        "用例编号", "用例类型", "功能模块", "测试目的", "前置条件", "优先级",
        "操作步骤", "输入数据", "预期结果",
        "执行状态", "执行结果", "对应缺陷ID", "编写人", "执行人", "备注",
    ]
    col_widths = [12, 10, 16, 24, 16, 10, 36, 28, 28, 10, 24, 16, 10, 10, 20]
    wb = Workbook()
    ws = wb.active
    ws.title = "用例列表"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, c in enumerate(cases, 2):
        created_by_name = names.get(c.created_by_id, "") if getattr(c, "created_by_id", None) else ""
        group_name = group_names.get(c.group_id, "") if c.group_id else ""
        case_type_val = c.type.value if hasattr(c.type, "value") else (c.type or "api")
        steps_str, input_data, expected_result = _format_export_columns(c.config or {}, case_type_val)
        ex = last_execution_map.get(c.id)
        exec_status = ""
        exec_result = ""
        executor_name = ""
        if ex:
            exec_status = _EXEC_STATUS_LABELS.get(
                ex.status.value if hasattr(ex.status, "value") else ex.status, str(ex.status)
            )
            if ex.logs:
                exec_result = (ex.logs or "")[:1000] + ("..." if len((ex.logs or "")) > 1000 else "")
            elif ex.result and isinstance(ex.result, dict) and ex.result.get("error"):
                exec_result = str(ex.result.get("error", ""))[:500]
            executor_name = executor_name_map.get(ex.created_by_id, "") if getattr(ex, "created_by_id", None) else ""
        defect_ids = defect_map.get(c.id, [])
        defect_id_str = ",".join(str(x) for x in defect_ids) if defect_ids else ""
        remark = ",".join(c.tags) if isinstance(c.tags, list) and c.tags else ""
        type_label = _TYPE_LABELS.get(case_type_val, case_type_val or "接口")
        row = [
            c.id,
            type_label,
            group_name,
            (c.description or "")[:2000],
            "",  # 前置条件暂无单独字段
            _PRIORITY_LABELS.get(c.priority.value if hasattr(c.priority, "value") else c.priority, c.priority),
            steps_str,
            input_data,
            expected_result,
            exec_status,
            exec_result,
            defect_id_str,
            created_by_name,
            executor_name,
            remark,
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in enumerate(col_widths, 1):
        if col <= len(col_widths):
            ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "cases_export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("", response_model=TestCaseOut)
async def create_case(data: TestCaseCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    payload = data.model_dump()
    payload["created_by_id"] = user.id
    collaborator_ids = payload.pop("collaborator_ids", None)
    if collaborator_ids is not None:
        payload["collaborator_ids"] = collaborator_ids
    group_id = payload.get("group_id")
    if group_id:
        gr = await db.execute(select(CaseGroup).where(CaseGroup.id == group_id))
        group = gr.scalar_one_or_none()
        if not group or group.project_id != payload.get("project_id"):
            raise HTTPException(400, "分组不存在或与项目不匹配")
        if not _can_manage_group(user, group):
            raise HTTPException(403, "仅该分组的创建人、管理员或协作者可在此分组下新建用例")
    case = TestCase(**payload)
    db.add(case)
    await db.commit()
    await db.refresh(case)
    return TestCaseOut.model_validate(case)


@router.get("/{case_id}", response_model=TestCaseOut)
async def get_case(case_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(404, "TestCase not found")
    return TestCaseOut.model_validate(case)


@router.put("/{case_id}", response_model=TestCaseOut)
async def update_case(case_id: int, data: TestCaseUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(404, "TestCase not found")
    if not _can_edit_case(user, case):
        raise HTTPException(403, "仅创建人、管理员或协作者可编辑")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(case, field, value)

    await db.commit()
    await db.refresh(case)
    return TestCaseOut.model_validate(case)


@router.delete("/{case_id}")
async def delete_case(case_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(404, "TestCase not found")
    if not _can_delete_case(user, case):
        raise HTTPException(403, "仅创建人或管理员可删除")

    await db.delete(case)
    await db.commit()
    return {"ok": True}


@router.post("/batch-delete")
async def batch_delete_cases(case_ids: list[int], db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    for cid in case_ids:
        result = await db.execute(select(TestCase).where(TestCase.id == cid))
        case = result.scalar_one_or_none()
        if case and _can_delete_case(user, case):
            await db.delete(case)
    await db.commit()
    return {"ok": True, "deleted": len(case_ids)}


@router.post("/batch-update")
async def batch_update_cases(data: TestCaseBatchUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """批量修改用例：优先级、状态、分组。仅对有编辑权限的用例生效。"""
    updated = 0
    for cid in data.case_ids:
        result = await db.execute(select(TestCase).where(TestCase.id == cid))
        case = result.scalar_one_or_none()
        if not case or not _can_edit_case(user, case):
            continue
        if data.priority is not None:
            case.priority = data.priority
        if data.status is not None:
            case.status = data.status
        if data.group_id is not None:
            case.group_id = data.group_id
        updated += 1
    await db.commit()
    return {"ok": True, "updated": updated}
